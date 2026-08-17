"""
10-generate-excel.py — 生成 Excel 评审汇总表

从每个 Subject 的 08-summarize 读取结构化评分，生成带 2 级合并表头的 Excel。
单篇论文也生成 Excel（一行数据），方便统一归档。

Excel 列结构（两级表头，第 1 列为论文名称固定列）：
         ┌────────┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┬──┐
         │        │      最终结果      │  间接维度打分   │ 原直接维度打分  │
         ├────────┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┼──┤
         │论文名称│创│质│效│风│难│业│总│行│问│公│源│业│前│创│质│效│风│难│业│
         │        │新│量│能│险│度│务│分│文│题│式│码│务│人│新│量│能│险│度│务│
         │        │性│提│提│敏│  │价│  │严│关│堆│深│规│调│性│提│提│敏│  │价│
         │        │  │升│升│感│  │值│  │谨│键│砌│度│模│研│  │升│升│感│  │值│
         │        │  │效│效│性│  │提│  │性│性│度│  │真│充│  │效│效│性│  │提│
         │        │  │果│果│  │  │升│  │  │  │  │  │实│分│  │果│果│  │  │升│
         │        │  │  │  │  │  │效│  │  │  │  │  │性│度│  │  │  │  │  │效│
         │        │  │  │  │  │  │果│  │  │  │  │  │  │  │  │  │  │  │  │果│
         └────────┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┴──┘

总分列：位于「最终结果」组末尾（业务价值提升效果之后），为 Excel 加权求和公式：
        =创新性*4+质量提升效果*3+效能提升效果*3+风险敏感性*3+难度*3+业务价值提升效果*2

依赖：openpyxl（pip install openpyxl）
"""

# pyright: reportPossiblyUnboundVariable=false, reportOptionalMemberAccess=false

from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    _HAS_OPENPYXL = False
else:
    _HAS_OPENPYXL = True

# ============================================================================
# 常量
# ============================================================================

DIRECT_DIMS = [
    "创新性",
    "质量提升效果",
    "效能提升效果",
    "风险敏感性",
    "难度",
    "业务价值提升效果",
]

INDIRECT_DIMS = [
    "行文严谨性",
    "问题关键性",
    "公式堆砌度",
    "源码深度",
    "业务规模真实性",
    "前人调研充分度",
]

# 总分权重：总分 = 创新*4 + 质量*3 + 效能*3 + 风险*3 + 难度*3 + 业务*2
TOTAL_WEIGHTS = {
    "创新性": 4,
    "质量提升效果": 3,
    "效能提升效果": 3,
    "风险敏感性": 3,
    "难度": 3,
    "业务价值提升效果": 2,
}

# 「最终结果」组末尾追加「总分」列（Excel 加权求和公式列）
FINAL_GROUP_DIMS = DIRECT_DIMS + ["总分"]

# 表头组配置
HEADER_GROUPS = [
    ("最终结果", FINAL_GROUP_DIMS),
    ("间接维度打分", INDIRECT_DIMS),
    ("原直接维度打分", DIRECT_DIMS),
    ("评分理由", ["评分理由"]),
]

# 样式（仅 openpyxl 可用时定义）
if _HAS_OPENPYXL:
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    SUB_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


# ============================================================================
# 构建 Excel
# ============================================================================


def _find_all_subjects(intermediates_dir: Path) -> list[str]:
    """扫描 intermediates 目录，找到所有有 08-summarize 产出的 Subject。"""
    subjects: list[str] = []
    if not intermediates_dir.exists():
        return subjects
    for subj_dir in sorted(intermediates_dir.iterdir()):
        if not subj_dir.is_dir():
            continue
        summ_path = subj_dir / "08-summarize" / "output.json"
        if summ_path.exists():
            subjects.append(subj_dir.name)
    return subjects


def _load_summarize(subject_name: str, intermediates_dir: Path) -> dict | None:
    """加载某个 Subject 的 08-summarize output.json，返回 data 字段。"""
    path = intermediates_dir / subject_name / "08-summarize" / "output.json"
    if not path.exists():
        return None
    # pi-lens-ignore: unchecked-throwing-call-python
    with open(path) as f:
        # pi-lens-ignore: unchecked-throwing-call-python
        raw = json.load(f)
    return raw.get("data", {}) if raw else None


def _auto_column_widths(ws, max_col: int):
    """自适应列宽。"""
    for col_idx in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 8)


def _total_formula(row_idx: int, final_start_col: int) -> str:
    """生成总分 Excel 公式：=创新*4+质量*3+效能*3+风险*3+难度*3+业务*2。"""
    terms = []
    for offset, dim in enumerate(DIRECT_DIMS):
        col = get_column_letter(final_start_col + offset)
        weight = TOTAL_WEIGHTS[dim]
        terms.append(f"{col}{row_idx}*{weight}")
    return "=" + "+".join(terms)


# ============================================================================
# 主逻辑
# ============================================================================


def main():
    step_dir = os.environ.get("PIPELINE_STEP_DIR", ".")
    output_dir = os.environ.get("PIPELINE_OUTPUT_DIR", ".")
    intermediates_dir = os.environ.get("PIPELINE_INTERMEDIATES", "")

    if not _HAS_OPENPYXL:
        print("10-generate-excel: openpyxl not installed — skipping Excel generation")
        output = {
            "step": "10-generate-excel",
            "status": "skipped",
            "error": "openpyxl not installed, Excel generation skipped",
            "data": {},
        }
        Path(step_dir).mkdir(parents=True, exist_ok=True)
        # pi-lens-ignore: unchecked-throwing-call-python
        with open(Path(step_dir) / "output.json", "w") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        return

    intermediates_path = Path(intermediates_dir)

    # ── 发现 Subject ──
    subjects = _find_all_subjects(intermediates_path)

    if len(subjects) == 0:
        print("10-generate-excel: no subjects — skipping Excel generation")
        output = {
            "step": "10-generate-excel",
            "status": "skipped",
            "error": "No subjects found, Excel not generated",
            "data": {"subject_count": 0},
        }
        Path(step_dir).mkdir(parents=True, exist_ok=True)
        # pi-lens-ignore: unchecked-throwing-call-python
        with open(Path(step_dir) / "output.json", "w") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        return

    # ── 构建 Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评审汇总"

    # 列布局：Col 1 = 论文名称（固定），Col 2+ = 维度数据
    total_dim_cols = sum(len(dims) for _, dims in HEADER_GROUPS)
    total_cols = 1 + total_dim_cols

    # Row 1: 一级表头 — Col 1 论文名称跨两行，Col 2+ 合并组
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    name_header = ws.cell(row=1, column=1, value="论文名称")
    name_header.font = HEADER_FONT
    name_header.fill = HEADER_FILL
    name_header.alignment = Alignment(horizontal="center", vertical="center")
    name_header.border = THIN_BORDER
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).border = THIN_BORDER

    dim_start = 2  # 维度列从第 2 列开始
    for group_name, dims in HEADER_GROUPS:
        end_col = dim_start + len(dims) - 1
        cell = ws.cell(row=1, column=dim_start, value=group_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if dim_start < end_col:
            ws.merge_cells(
                start_row=1,
                start_column=dim_start,
                end_row=1,
                end_column=end_col,
            )
        for c in range(dim_start, end_col + 1):
            ws.cell(row=1, column=c).fill = HEADER_FILL
            ws.cell(row=1, column=c).border = THIN_BORDER
        dim_start = end_col + 1

    # Row 2: 二级表头（子维度名，从 Col 2 开始）
    col_idx = 2
    for group_name, dims in HEADER_GROUPS:
        for dim in dims:
            cell = ws.cell(row=2, column=col_idx, value=dim)
            cell.font = Font(bold=True, size=10)
            cell.fill = SUB_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            col_idx += 1

    # Data rows: 每个 Subject 一行
    for row_idx, subject in enumerate(subjects, start=3):
        # Col 1: Subject 名称
        name_cell = ws.cell(row=row_idx, column=1, value=subject)
        name_cell.font = Font(bold=True)
        name_cell.border = THIN_BORDER

        data = _load_summarize(subject, intermediates_path)
        if not data:
            for c in range(2, total_cols + 1):
                ws.cell(row=row_idx, column=c, value="N/A").border = THIN_BORDER
            continue

        final_scores = data.get("final_scores", {})
        indirect_scores = data.get("indirect_scores", {})
        original_scores = data.get("original_direct_scores", {})

        col_idx = 2
        # 最终结果（直接维度 + 总分公式）
        final_start_col = col_idx
        for dim in DIRECT_DIMS:
            ws.cell(
                row=row_idx, column=col_idx, value=final_scores.get(dim, "-")
            ).border = THIN_BORDER
            col_idx += 1
        # 总分（Excel 加权求和公式）
        ws.cell(
            row=row_idx, column=col_idx, value=_total_formula(row_idx, final_start_col)
        ).border = THIN_BORDER
        col_idx += 1
        # 间接维度
        for dim in INDIRECT_DIMS:
            ws.cell(
                row=row_idx, column=col_idx, value=indirect_scores.get(dim, "-")
            ).border = THIN_BORDER
            col_idx += 1
        # 原直接维度
        for dim in DIRECT_DIMS:
            ws.cell(
                row=row_idx, column=col_idx, value=original_scores.get(dim, "-")
            ).border = THIN_BORDER
            col_idx += 1
        # 评分理由（最后一列：逐维 rationale + 降级原因）
        rationale = data.get("rationale", {})
        degradation_reason = data.get("degradation_reason", "")
        reason_lines = []
        for dim in DIRECT_DIMS:
            r = (rationale or {}).get(dim)
            if r:
                reason_lines.append(f"{dim}：{r}")
        if degradation_reason:
            reason_lines.append(degradation_reason)
        reason_cell = ws.cell(row=row_idx, column=col_idx, value="\n".join(reason_lines))
        reason_cell.border = THIN_BORDER
        reason_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        col_idx += 1

    # 列宽自适应
    _auto_column_widths(ws, total_cols)
    # 评分理由列固定宽度（长文本换行，不参与自适应撑爆）
    ws.column_dimensions[get_column_letter(total_cols)].width = 60

    # ── 写文件 ──
    output_path = Path(output_dir)
    reports_dir = output_path / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    excel_path = reports_dir / f"summary-{timestamp}.xlsx"

    wb.save(str(excel_path))

    print(f"10-generate-excel: saved to {excel_path}")
    print(f"  {len(subjects)} subjects, {total_cols} columns")

    # ── 写 output.json ──
    output = {
        "step": "10-generate-excel",
        "status": "ok",
        "error": None,
        "data": {
            "excel_path": str(excel_path),
            "subject_count": len(subjects),
            "column_count": total_cols,
        },
    }

    Path(step_dir).mkdir(parents=True, exist_ok=True)
    # pi-lens-ignore: unchecked-throwing-call-python
    with open(Path(step_dir) / "output.json", "w") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
