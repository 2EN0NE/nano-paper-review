"""08-summarize.py 管线模板测试 —— rationale/tags 提取 + 字段级降级标记。

验证：
  1. 结构化 JSON 输入 → rationale/tags 正确提取，无降级。
  2. raw_output markdown 输入 → rationale 从表格第三列提取（降级兜底）。
  3. rationale/tags 缺失 → evidence 字段级标记 + degradation_reason。
"""

from __future__ import annotations

import importlib.util
import json
import os
import runpy
from pathlib import Path

import pytest

SUMMARIZE_TEMPLATE = (
    Path(__file__).resolve().parent.parent
    / "src/paper_review/templates/review-pipeline/08-summarize.py"
)
EXCEL_TEMPLATE = (
    Path(__file__).resolve().parent.parent
    / "src/paper_review/templates/post-review/10-generate-excel.py"
)

DIRECT_DIMS = ["创新性", "质量提升效果", "效能提升效果", "风险敏感性", "难度", "业务价值提升效果"]
INDIRECT_DIMS = [
    "行文严谨性",
    "问题关键性",
    "公式堆砌度",
    "源码深度",
    "业务规模真实性",
    "前人调研充分度",
]


def _load_module():
    spec = importlib.util.spec_from_file_location("summarize", SUMMARIZE_TEMPLATE)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _run_summarize(env: dict) -> dict:
    step_dir = Path(env["PIPELINE_STEP_DIR"])
    step_dir.mkdir(parents=True, exist_ok=True)
    old = {k: os.environ.get(k) for k in env}
    try:
        os.environ.update(env)
        runpy.run_path(str(SUMMARIZE_TEMPLATE), run_name="__main__")
    finally:
        for k, v in old.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
    return json.loads((step_dir / "output.json").read_text(encoding="utf-8"))


def _write_scoring(intermediates: Path, subject: str, step: str, data: dict) -> None:
    d = intermediates / subject / step
    d.mkdir(parents=True, exist_ok=True)
    (d / "output.json").write_text(
        json.dumps({"step": step, "status": "ok", "data": data}, ensure_ascii=False),
        encoding="utf-8",
    )


# ============================================================================
# 纯函数测试
# ============================================================================


class TestExtractRationaleFromMarkdown:
    def test_third_column_extracted(self):
        mod = _load_module()
        md = (
            "| 维度 | 分数 | 核心依据 |\n"
            "|------|------|----------|\n"
            "| 创新性 | 5 | KDD Cup 冠军（1/1895） |\n"
            "| 难度 | 4 | 攻克部门内难题 |\n"
        )
        rationale = mod._extract_rationale_from_markdown(md, DIRECT_DIMS)
        assert rationale["创新性"] == "KDD Cup 冠军（1/1895）"
        assert rationale["难度"] == "攻克部门内难题"
        # 表头行「维度」不在 dims 中，被过滤
        assert "维度" not in rationale

    def test_empty_reason_skipped(self):
        mod = _load_module()
        md = "| 创新性 | 5 |  |\n"
        assert mod._extract_rationale_from_markdown(md, DIRECT_DIMS) == {}


class TestParseScoringRationale:
    def test_structured_dict(self):
        mod = _load_module()
        data = {"创新性": {"score": 5, "rationale": "冠军证据"}}
        assert mod._parse_scoring_rationale(data, DIRECT_DIMS) == {"创新性": "冠军证据"}

    def test_raw_output_fallback(self):
        mod = _load_module()
        data = {"raw_output": "| 创新性 | 5 | 冠军证据 |\n"}
        assert mod._parse_scoring_rationale(data, DIRECT_DIMS) == {"创新性": "冠军证据"}

    def test_empty_rationale_skipped(self):
        mod = _load_module()
        data = {"创新性": {"score": 5, "rationale": "  "}}
        assert mod._parse_scoring_rationale(data, DIRECT_DIMS) == {}


class TestParseTags:
    def test_list(self):
        mod = _load_module()
        assert mod._parse_tags({"tags": ["去偏", "i2i"]}) == ["去偏", "i2i"]

    def test_missing_or_invalid(self):
        mod = _load_module()
        assert mod._parse_tags({}) == []
        assert mod._parse_tags({"tags": "not-a-list"}) == []

    def test_filters_non_string(self):
        mod = _load_module()
        assert mod._parse_tags({"tags": ["去偏", 123, "", "  "]}) == ["去偏"]


# ============================================================================
# main 流程测试
# ============================================================================


class TestSummarizeMain:
    def _env(self, tmp_path: Path, intermediates: Path, subject: str) -> dict:
        return {
            "PIPELINE_STEP_DIR": str(tmp_path / "step"),
            "PIPELINE_INTERMEDIATES": str(intermediates),
            "PIPELINE_SUBJECT": subject,
        }

    def test_structured_input_no_degradation(self, tmp_path):
        """结构化 direct（含 rationale+tags）→ 无降级，rationale/tags 进输出。"""
        intermediates = tmp_path / "intermediates"
        _write_scoring(
            intermediates,
            "论文A",
            "06-direct-scoring",
            {
                "tags": ["去偏", "i2i", "流行度惩罚"],
                **{d: {"score": 4, "rationale": f"{d}理由"} for d in DIRECT_DIMS},
            },
        )
        _write_scoring(
            intermediates,
            "论文A",
            "07-indirect-scoring",
            {d: {"score": 3, "rationale": f"{d}理由"} for d in INDIRECT_DIMS},
        )

        data = _run_summarize(self._env(tmp_path, intermediates, "论文A"))["data"]
        assert data["degraded"] is False
        assert data["degradation_reason"] == ""
        assert data["tags"] == ["去偏", "i2i", "流行度惩罚"]
        assert data["rationale"]["创新性"] == "创新性理由"
        assert data["evidence"] == {"rationale_missing": [], "tags_missing": False}

    def test_raw_output_markdown_tags_degraded(self, tmp_path):
        """direct 只有 raw_output markdown（无 tags）→ rationale 从表格抠，tags 降级。"""
        intermediates = tmp_path / "intermediates"
        md = "已完成打分\n| 维度 | 分数 | 核心依据 |\n|------|------|----------|\n" + "".join(
            f"| {d} | 3 | {d}依据 |\n" for d in DIRECT_DIMS
        )
        _write_scoring(intermediates, "论文B", "06-direct-scoring", {"raw_output": md})
        _write_scoring(
            intermediates,
            "论文B",
            "07-indirect-scoring",
            {d: {"score": 3, "rationale": f"{d}理由"} for d in INDIRECT_DIMS},
        )

        data = _run_summarize(self._env(tmp_path, intermediates, "论文B"))["data"]
        # rationale 从 markdown 第三列抠到 → 无 rationale_missing
        assert data["rationale"]["创新性"] == "创新性依据"
        assert data["evidence"]["rationale_missing"] == []
        # tags 缺失 → tags_missing + degraded
        assert data["evidence"]["tags_missing"] is True
        assert data["degraded"] is True
        assert "tags缺失" in data["degradation_reason"]

    def test_skip_article_all_rationale_missing(self, tmp_path):
        """direct raw_output 无表格（软文跳过）→ 分数默认 3.0，全维度 rationale 缺失。"""
        intermediates = tmp_path / "intermediates"
        _write_scoring(
            intermediates,
            "论文C",
            "06-direct-scoring",
            {"raw_output": "本文是软文，跳过打分。"},
        )
        _write_scoring(
            intermediates,
            "论文C",
            "07-indirect-scoring",
            {d: {"score": 3, "rationale": f"{d}理由"} for d in INDIRECT_DIMS},
        )

        data = _run_summarize(self._env(tmp_path, intermediates, "论文C"))["data"]
        assert data["original_direct_scores"]["创新性"] == 3.0
        assert set(data["evidence"]["rationale_missing"]) == set(DIRECT_DIMS)
        assert "缺证据" in data["degradation_reason"]


# ============================================================================
# 10-generate-excel — 评分理由列
# ============================================================================


def _run_excel(env: dict) -> dict:
    step_dir = Path(env["PIPELINE_STEP_DIR"])
    step_dir.mkdir(parents=True, exist_ok=True)
    old = {k: os.environ.get(k) for k in env}
    try:
        os.environ.update(env)
        runpy.run_path(str(EXCEL_TEMPLATE), run_name="__main__")
    finally:
        for k, v in old.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
    return json.loads((step_dir / "output.json").read_text(encoding="utf-8"))


def _load_excel_module():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not installed")
    spec = importlib.util.spec_from_file_location("generate_excel", EXCEL_TEMPLATE)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


class TestGenerateExcelRatingReason:
    def test_rating_reason_column(self, tmp_path):
        """10-generate-excel 生成 21 列，最后一列含逐维 rationale + 降级原因。"""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        intermediates = tmp_path / "intermediates"
        output_dir = tmp_path / "output"
        output_dir.mkdir(parents=True)
        summ_dir = intermediates / "论文A" / "08-summarize"
        summ_dir.mkdir(parents=True, exist_ok=True)
        (summ_dir / "output.json").write_text(
            json.dumps(
                {
                    "step": "08-summarize",
                    "status": "ok",
                    "data": {
                        "final_scores": {d: 3.0 for d in DIRECT_DIMS},
                        "indirect_scores": {d: 3.0 for d in INDIRECT_DIMS},
                        "original_direct_scores": {d: 3.0 for d in DIRECT_DIMS},
                        "rationale": {d: f"{d}理由" for d in DIRECT_DIMS},
                        "tags": ["去偏"],
                        "degraded": True,
                        "degradation_reason": "⚠ 缺证据:难度",
                    },
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        out = _run_excel(
            {
                "PIPELINE_STEP_DIR": str(tmp_path / "step"),
                "PIPELINE_OUTPUT_DIR": str(output_dir),
                "PIPELINE_INTERMEDIATES": str(intermediates),
            }
        )
        assert out["data"]["subject_count"] == 1
        assert out["data"]["column_count"] == 21

        wb = openpyxl.load_workbook(out["data"]["excel_path"])
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=2, column=21).value == "评分理由"
        reason = str(ws.cell(row=3, column=21).value)
        assert "创新性：创新性理由" in reason
        assert "⚠ 缺证据:难度" in reason

    def test_total_formula_column(self, tmp_path):
        """总分列（第 8 列）写入加权求和公式，引用前 6 个直接维度列。"""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        intermediates = tmp_path / "intermediates"
        output_dir = tmp_path / "output"
        output_dir.mkdir(parents=True)
        summ_dir = intermediates / "论文A" / "08-summarize"
        summ_dir.mkdir(parents=True, exist_ok=True)
        (summ_dir / "output.json").write_text(
            json.dumps(
                {
                    "step": "08-summarize",
                    "status": "ok",
                    "data": {
                        "final_scores": {d: 3.0 for d in DIRECT_DIMS},
                        "indirect_scores": {d: 3.0 for d in INDIRECT_DIMS},
                        "original_direct_scores": {d: 3.0 for d in DIRECT_DIMS},
                        "rationale": {},
                        "tags": [],
                    },
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        out = _run_excel(
            {
                "PIPELINE_STEP_DIR": str(tmp_path / "step"),
                "PIPELINE_OUTPUT_DIR": str(output_dir),
                "PIPELINE_INTERMEDIATES": str(intermediates),
            }
        )

        wb = openpyxl.load_workbook(out["data"]["excel_path"])
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=2, column=8).value == "总分"
        # final_start_col=2 → 维度列 B..G，权重 4/3/3/3/3/2
        assert ws.cell(row=3, column=8).value == "=B3*4+C3*3+D3*3+E3*3+F3*3+G3*2"


class TestTotalFormula:
    """10-generate-excel 总分加权公式（业务规则，纯函数）。"""

    def test_total_weights(self):
        mod = _load_excel_module()
        assert mod.TOTAL_WEIGHTS == {
            "创新性": 4,
            "质量提升效果": 3,
            "效能提升效果": 3,
            "风险敏感性": 3,
            "难度": 3,
            "业务价值提升效果": 2,
        }

    def test_formula_string(self):
        mod = _load_excel_module()
        # final_start_col=2 → 维度列 B..G，权重 4/3/3/3/3/2
        assert mod._total_formula(3, 2) == "=B3*4+C3*3+D3*3+E3*3+F3*3+G3*2"

    def test_formula_uses_column_letters(self):
        mod = _load_excel_module()
        # 数据行与维度组起始列偏移时，列字母随列号变化
        assert mod._total_formula(5, 10) == "=J5*4+K5*3+L5*3+M5*3+N5*3+O5*2"
